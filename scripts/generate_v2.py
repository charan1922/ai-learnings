"""Generate IT_Tickets_v2.xlsx: copies all sheets from v1, adds Access Management sheet (200 rows)."""

import random
import openpyxl

SRC = "docs/IT_Tickets_v1.xlsx"
DST = "docs/IT_Tickets_v2.xlsx"

SYSTEMS = [
    "Active Directory",
    "LDAP Server",
    "SSO Portal",
    "IAM System",
    "Role Management System",
    "User Provisioning Service",
    "Password Reset Portal",
    "OAuth Server",
    "Privileged Access Manager",
    "Service Account Manager",
]

ISSUES = [
    ("Unauthorized Access Attempt", "Unauthorized access attempt detected on {system}."),
    ("Permission Escalation", "User escalated permissions without authorization on {system}."),
    ("Account Lockout", "Multiple failed logins caused account lockout on {system}."),
    ("Role Misconfiguration", "Role misconfiguration detected on {system} causing overprivileged access."),
    ("Expired Credentials", "Service disruption on {system} due to expired credentials."),
    ("Access Policy Violation", "Access policy violation flagged on {system}."),
    ("Privileged Account Abuse", "Privileged account abuse detected on {system}. Immediate review required."),
    ("Service Account Compromise", "Indicators of compromise found on service account in {system}."),
    ("Failed SSO Authentication", "Multiple users reporting failed SSO authentication against {system}."),
    ("Account Provisioning Failure", "New user account provisioning failed on {system}."),
    ("MFA Bypass Attempt", "Suspicious MFA bypass attempt recorded on {system}."),
    ("Access Token Expiry", "Access tokens expired prematurely on {system} blocking user sessions."),
    ("Group Policy Conflict", "Group policy conflict causing inconsistent permissions on {system}."),
    ("Certificate Expiry", "Authentication certificate expired on {system} causing login failures."),
    ("Inactive Account Access", "Dormant account used to access {system} outside business hours."),
    ("Stale Permission Review", "Scheduled review found stale permissions accumulating on {system}."),
    ("Directory Sync Failure", "Directory synchronization failure between {system} and downstream apps."),
    ("Federation Trust Error", "Federation trust configuration error on {system} blocking external logins."),
]

RESOLUTIONS = [
    "Revoked access and issued new credentials.",
    "Updated user roles and removed excess permissions.",
    "Reset account and enforced MFA policy.",
    "Applied least-privilege policy and reviewed ACLs.",
    "Deactivated stale accounts and notified user managers.",
    "Rotated service account credentials and updated vaults.",
    "Patched LDAP configuration and restarted service.",
    "Reconfigured SSO federation settings.",
    "Reviewed and updated ACL rules across affected groups.",
    "Enforced password complexity and rotation policy.",
    "Blocked unauthorized IP and reset compromised credentials.",
    "Removed excessive permissions and filed access review ticket.",
    "Rebuilt group policy objects and re-applied to OU.",
    "Renewed authentication certificates and restarted service.",
    "Disabled compromised service account and rotated secrets.",
    "Provisioned replacement account and verified access.",
    "Resolved directory sync conflict and validated user data.",
    "Re-established federation trust and tested external login.",
]

PRIORITIES = ["Low", "Medium", "High", "Critical"]
PRIORITY_WEIGHTS = [0.25, 0.40, 0.25, 0.10]

DESCRIPTION_PREFIXES = [
    "User reported {issue_lower} on the {system}. It is currently affecting operations.",
    "Support requested for {system} due to {issue_lower}.",
    "Automated alert: {system} flagged {issue_lower}. Immediate investigation required.",
    "Diagnostics indicate {issue_lower} related to {system}.",
    "Security team escalated {issue_lower} detected on {system}.",
]


def random_ticket_id(existing: set) -> str:
    while True:
        tid = f"TKT-{random.randint(100000, 999999)}"
        if tid not in existing:
            existing.add(tid)
            return tid


def build_access_management_rows(n: int = 200) -> list[tuple]:
    random.seed(42)
    used_ids: set = set()
    rows = []
    for _ in range(n):
        system = random.choice(SYSTEMS)
        issue_title, issue_desc_template = random.choice(ISSUES)
        title = f"{system} - {issue_title}"
        desc_template = random.choice(DESCRIPTION_PREFIXES)
        description = desc_template.format(
            system=system,
            issue_lower=issue_title.lower(),
        )
        resolution = random.choice(RESOLUTIONS)
        priority = random.choices(PRIORITIES, weights=PRIORITY_WEIGHTS, k=1)[0]
        ticket_id = random_ticket_id(used_ids)
        rows.append((ticket_id, title, description, "Access Management", resolution, priority))
    return rows


def main() -> None:
    src_wb = openpyxl.load_workbook(SRC)
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)

    # Copy existing sheets
    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        dst_ws = dst_wb.create_sheet(title=sheet_name)
        for row in src_ws.iter_rows(values_only=True):
            dst_ws.append(list(row))
        print(f"Copied sheet: {sheet_name} ({src_ws.max_row - 1} data rows)")

    # Add Access Management sheet
    am_ws = dst_wb.create_sheet(title="Access Management")
    headers = ["Ticket ID", "Title", "Description", "Category", "Resolution", "Priority"]
    am_ws.append(headers)
    for row in build_access_management_rows(200):
        am_ws.append(list(row))
    print(f"Created sheet: Access Management (200 data rows)")

    dst_wb.save(DST)
    print(f"\nSaved: {DST}")
    print(f"Sheets: {dst_wb.sheetnames}")


if __name__ == "__main__":
    main()
